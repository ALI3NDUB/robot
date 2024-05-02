import pandas as pd
import openpyxl
from io import BytesIO
import streamlit as st  

def process_excel(df_csv_l1, df_csv_l0, cliente, Nome_progetto, campi_tecnici):
    wb = openpyxl.Workbook()

    # Elimina il foglio predefinito
    wb.remove(wb.active)

    # Aggiungi fogli con nomi specifici
    wb.create_sheet("Mapping Data")
    wb.create_sheet("Properties")
    wb.create_sheet("Readme")

    ws_mapping = wb["Mapping Data"]
    ws_properties = wb["Properties"]
    ws_readme = wb["Readme"]

    # Foglio Mapping Data
    headers_mapping = ['MAPPING_NAME', 'COMPONENT_NAME', 'TARGET_TABLE', 'MAP_GROUP_NAME', 'SET_OPERATION',
               'TARGET_COLUMN', 'MAP_EXPRESSION', 'TARGET_COLUMN_FLAGS', 'SOURCE_TABLE', 'FILTER_CONDITION',
               'GROUP_BY', 'ORDER_BY', 'DISTINCT', 'TARGET_CONDITION']
    for col, header in enumerate(headers_mapping, start=1):
        ws_mapping.cell(row=1, column=col, value=header)

    excel_row_counter_mapping = 2
    target_table_inserted = {}
    mapping_name_inserted = {}
    for _, row in df_csv_l1.iterrows():
        if row['COLUMN_NAME'] not in campi_tecnici:
            target_table = row['OWNER'] + '.' + row['TABLE_NAME']
            mapping_name = f"{cliente}.{Nome_progetto}.L1_{row['TABLE_NAME']}"
            if mapping_name not in mapping_name_inserted:
                ws_mapping.cell(row=excel_row_counter_mapping, column=1, value=mapping_name)
                mapping_name_inserted[mapping_name] = True
            if target_table not in target_table_inserted:
                ws_mapping.cell(row=excel_row_counter_mapping, column=3, value=target_table)
                target_table_inserted[target_table] = True
            ws_mapping.cell(row=excel_row_counter_mapping, column=6, value=row['COLUMN_NAME'])
            corresponding_row = df_csv_l0[(df_csv_l0['TABLE_NAME'] == row['TABLE_NAME'].replace('OK', 'DLT')) &
                                          (df_csv_l0['COLUMN_NAME'] == row['COLUMN_NAME'])]
            if not corresponding_row.empty:
                map_expression = f"SRC.{corresponding_row.iloc[0]['COLUMN_NAME']}"
                ws_mapping.cell(row=excel_row_counter_mapping, column=7, value=map_expression)
            if not df_csv_l0.empty and any(df_csv_l0['TABLE_NAME'] == row['TABLE_NAME'].replace('OK', 'DLT')):
                source_table = f"{row['OWNER']}.{row['TABLE_NAME'].replace('OK', 'DLT')} SRC"
                ws_mapping.cell(row=excel_row_counter_mapping, column=9, value=source_table)
            excel_row_counter_mapping += 1

    # Foglio Properties
    # Aggiungi qui il codice per il foglio Properties...

    # Foglio Readme
    # Aggiungi qui il codice per il foglio Readme...

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()

def main():
    st.title("Processo Excel da CSV")
    csv_file = st.file_uploader("Carica CSV", type=["csv"])
    if csv_file is not None:
        st.subheader("Anteprima CSV:")
        df_csv_preview = pd.read_csv(csv_file)
        st.dataframe(df_csv_preview.head())
        df_csv_l1 = df_csv_preview[df_csv_preview['TABLE_NAME'].str.startswith('OK')]
        df_csv_l0 = df_csv_preview[df_csv_preview['TABLE_NAME'].str.startswith('DLT')]
        cliente = st.text_input("Cliente")
        nome_progetto = st.text_input("Nome Progetto")
        campi_tecnici = st.text_input("Campi Tecnici (separati da virgola)").split(',')
        if st.button("Esegui Processo"):
            processed_excel_data = process_excel(df_csv_l1, df_csv_l0, cliente, nome_progetto, campi_tecnici)
            st.write("Processo completato!")
            st.download_button(
                label="Scarica Excel Processato",
                data=processed_excel_data,
                file_name="excel_process.xlsx"  # Corretto il nome del file
            )

if __name__ == "__main__":
    main()
